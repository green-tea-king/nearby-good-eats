import html
import io
import json
import re
import sys
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook


SOURCE_PAGE_URL = "https://www.fda.gov.tw/tc/siteContent.aspx?sid=13237"
OUT_PATH = r"C:\Users\Administrator\Documents\Codex\2026-06-21\github\nearby-good-eats\assets\fda-restaurant-hygiene-2024-candidates.json"
REPORT_PATH = r"C:\Users\Administrator\Documents\Codex\2026-06-21\github\nearby-good-eats\assets\fda-restaurant-hygiene-2024-import-report.json"
ROC_YEAR = 113
WESTERN_YEAR = 2024
GUIDE = "fdagrade"
AWARD_NAME = "餐飲衛生管理分級評核"
AUTHORITY = "衛生福利部食品藥物管理署"

CITY_PATTERN = re.compile(r"^(臺北市|新北市|桃園市|臺中市|臺南市|高雄市|基隆市|新竹市|嘉義市|新竹縣|苗栗縣|彰化縣|南投縣|雲林縣|嘉義縣|屏東縣|宜蘭縣|花蓮縣|臺東縣|澎湖縣|金門縣|連江縣)")
DISTRICT_PATTERN = re.compile(r"(.*?(?:區|鄉|鎮|市))")
LINK_PATTERN = re.compile(r'href="([^"]+GetFile\.ashx\?id=[^"]+type=4)"[^>]*>([^<]*\(excel\)[^<]*)<', re.I)


def taipei_now_iso():
    return datetime.now(timezone(timedelta(hours=8))).isoformat()


def fetch_text(url):
    with urllib.request.urlopen(url, timeout=120) as res:
        return res.read().decode("utf-8", errors="replace")


def fetch_bytes(url):
    with urllib.request.urlopen(url, timeout=240) as res:
        return res.read()


def normalize_space(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_city(address):
    matched = CITY_PATTERN.match(address or "")
    return matched.group(1) if matched else ""


def parse_district(address, city):
    rest = address[len(city):] if city and address.startswith(city) else address
    matched = DISTRICT_PATTERN.match(rest or "")
    return matched.group(1) if matched else ""


def resolve_url(href):
    href = href.replace("&amp;", "&")
    return href if href.startswith("http") else f"https://www.fda.gov.tw{href}"


def detect_columns(header):
    columns = {}
    for idx, cell in enumerate(header):
        value = normalize_space(cell)
        if not value:
            continue
        if "店名" in value:
            columns["name"] = idx
        elif "地址" in value:
            columns["address"] = idx
        elif "餐飲業別" in value:
            columns["cuisine"] = idx
        elif "評核等級" in value:
            columns["level"] = idx
        elif "電話" in value:
            columns["phone"] = idx
        elif "食品業者登錄字號" in value:
            columns["license"] = idx
    return columns


def parse_workbook(city, source_url, blob, extracted_at):
    workbook = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    columns = {}
    restaurants = []
    skipped = 0
    header_found = False

    for raw in sheet.iter_rows(values_only=True):
        values = [normalize_space(v) for v in raw]
        non_empty = [v for v in values if v]
        if not non_empty:
            continue
        if not header_found:
            if any("店名" in v for v in non_empty) and any("評核等級" in v for v in non_empty):
                columns = detect_columns(values)
                header_found = True
            continue

        name = values[columns.get("name", -1)] if "name" in columns else ""
        address = values[columns.get("address", -1)] if "address" in columns else ""
        cuisine = values[columns.get("cuisine", -1)] if "cuisine" in columns else ""
        level = values[columns.get("level", -1)] if "level" in columns else ""
        phone = values[columns.get("phone", -1)] if "phone" in columns else ""
        license_no = values[columns.get("license", -1)] if "license" in columns else ""

        if not name or not address or level not in ("優", "良"):
            skipped += 1
            continue

        parsed_city = parse_city(address) or city
        district = parse_district(address, parsed_city)
        notes = [f"來源檔案標示 {city}{ROC_YEAR}年度通過餐飲衛生管理分級評核名單。"]
        if phone:
            notes.append(f"電話 {phone}")
        if license_no:
            notes.append(f"登錄字號 {license_no}")

        restaurants.append({
            "name": name,
            "city": parsed_city,
            "district": district,
            "address": address,
            "cuisine": cuisine or "餐飲業者",
            "aliases": [],
            "awards": [{
                "guide": GUIDE,
                "year": WESTERN_YEAR,
                "level": level,
                "awardName": AWARD_NAME,
                "sourceName": AUTHORITY,
                "url": source_url,
                "extractedAt": extracted_at,
                "notes": " ".join(notes),
            }],
            "importConfidence": "high",
            "matchedBy": "official_city_excel_row",
        })

    return restaurants, skipped, sorted(columns.keys())


def main():
    generated_at = taipei_now_iso()
    report = {
        "generatedAt": generated_at,
        "sourcePageUrl": SOURCE_PAGE_URL,
        "sourceFiles": 0,
        "restaurants": 0,
        "cities": 0,
        "skippedRows": 0,
        "errors": [],
    }
    try:
        page = fetch_text(SOURCE_PAGE_URL)
        files = []
        seen = set()
        for href, text in LINK_PATTERN.findall(page):
            label = html.unescape(text).strip()
            city = label.split(str(ROC_YEAR))[0]
            if city in seen:
                continue
            seen.add(city)
            files.append({"city": city, "label": label, "url": resolve_url(href)})

        restaurants = []
        source_files = []
        for item in files:
            rows, skipped, columns = parse_workbook(item["city"], item["url"], fetch_bytes(item["url"]), generated_at)
            restaurants.extend(rows)
            report["skippedRows"] += skipped
            source_files.append({
                "city": item["city"],
                "label": item["label"],
                "url": item["url"],
                "rows": len(rows),
                "columns": columns,
            })

        payload = {
            "version": "fda-restaurant-hygiene-2024-candidates",
            "generatedAt": generated_at,
            "sourcePageUrl": SOURCE_PAGE_URL,
            "sourceFiles": source_files,
            "policy": {
                "runtimeExternalLookup": False,
                "batchOnly": True,
                "importMode": "official_excel_batch_import",
                "notes": [
                    "來源為食藥署 113年度各縣市衛生優良餐飲業者名單 excel 附件。",
                    "ROC 113 年換算西元 2024 年。",
                    "資料逐列取自官方附件，不使用即時 API。",
                ],
            },
            "restaurants": restaurants,
        }

        report["sourceFiles"] = len(source_files)
        report["restaurants"] = len(restaurants)
        report["cities"] = len({row["city"] for row in restaurants})

        with open(OUT_PATH, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")
        with open(REPORT_PATH, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(json.dumps(report, ensure_ascii=False, indent=2))
    except Exception as exc:
        report["errors"].append(str(exc))
        with open(REPORT_PATH, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(json.dumps(report, ensure_ascii=False, indent=2), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
